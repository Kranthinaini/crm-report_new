import math
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Page setup
# ============================================================

st.set_page_config(
    page_title="Field Visit Compliance",
    page_icon="🧭",
    layout="wide",
)

# ============================================================
# Visual identity
# ============================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;600&display=swap');

:root {
    --navy: #12283B;
    --navy-2: #1C3A52;
    --bg: #F5F7F8;
    --card: #FFFFFF;
    --border: #E1E6EA;
    --ink: #1B2733;
    --muted: #64707A;
    --teal: #1C9C8B;
    --teal-bg: #E1F5F1;
    --amber: #E8A33D;
    --amber-bg: #FCF1DF;
    --coral: #E2584A;
    --coral-bg: #FBE7E4;
}

html, body, [class*="css"]  {
    font-family: 'Inter', sans-serif;
    color: var(--ink);
}

.stApp {
    background: var(--bg);
}

#MainMenu, footer, header {visibility: hidden;}

.block-container {
    padding-top: 2rem;
    max-width: 1180px;
}

/* ---------- Hero ---------- */

.hero {
    background: linear-gradient(135deg, var(--navy) 0%, var(--navy-2) 100%);
    border-radius: 18px;
    padding: 2.4rem 2.6rem;
    margin-bottom: 1.8rem;
    position: relative;
    overflow: hidden;
}

.hero-eyebrow {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.75rem;
    letter-spacing: 0.14em;
    color: #7FD9CC;
    text-transform: uppercase;
    margin-bottom: 0.6rem;
}

.hero h1 {
    font-family: 'Sora', sans-serif;
    font-weight: 800;
    font-size: 2.1rem;
    color: #FFFFFF;
    margin: 0 0 0.5rem 0;
    letter-spacing: -0.02em;
}

.hero p {
    color: #B9C6D2;
    font-size: 0.98rem;
    margin: 0;
    max-width: 640px;
}

.route-line {
    height: 2px;
    margin-top: 1.6rem;
    background: repeating-linear-gradient(to right, #7FD9CC 0 9px, transparent 9px 18px);
    position: relative;
}
.route-line::before, .route-line::after {
    content: '';
    position: absolute;
    top: 50%;
    transform: translateY(-50%);
    width: 8px; height: 8px;
    border-radius: 50%;
    background: #7FD9CC;
}
.route-line::before { left: -4px; }
.route-line::after { right: -4px; }

/* ---------- Section divider (signature element) ---------- */

.checkpoint-divider {
    height: 2px;
    margin: 2.2rem 0;
    background: repeating-linear-gradient(to right, var(--teal) 0 9px, transparent 9px 20px);
    position: relative;
}
.checkpoint-divider::before, .checkpoint-divider::after {
    content: '';
    position: absolute;
    top: 50%;
    transform: translateY(-50%);
    width: 9px; height: 9px;
    border-radius: 50%;
    background: var(--navy);
}
.checkpoint-divider::before { left: -4px; }
.checkpoint-divider::after { right: -4px; }

/* ---------- Cards ---------- */

.card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.4rem 1.6rem;
}

.section-label {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 0.6rem;
}

/* ---------- KPI cards ---------- */

.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.9rem;
    margin-bottom: 0.4rem;
}

.kpi-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-left: 4px solid var(--navy);
    border-radius: 12px;
    padding: 1.1rem 1.3rem;
}

.kpi-card.total { border-left-color: var(--navy); }
.kpi-card.full { border-left-color: var(--teal); }
.kpi-card.partial { border-left-color: var(--amber); }
.kpi-card.non { border-left-color: var(--coral); }

.kpi-label {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 0.35rem;
}

.kpi-value {
    font-family: 'Sora', sans-serif;
    font-weight: 700;
    font-size: 1.9rem;
    color: var(--ink);
    line-height: 1;
}

.kpi-sub {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem;
    color: var(--muted);
    margin-top: 0.3rem;
}

/* ---------- Uploader ---------- */

[data-testid="stFileUploaderDropzone"] {
    background: var(--card);
    border: 1.5px dashed #C4CDD3;
    border-radius: 14px;
}

/* ---------- Buttons ---------- */

.stDownloadButton button, .stButton button {
    background: var(--navy);
    color: #FFFFFF;
    border: none;
    border-radius: 10px;
    padding: 0.6rem 1.4rem;
    font-family: 'Inter', sans-serif;
    font-weight: 600;
    letter-spacing: 0.01em;
    transition: background 0.15s ease;
}
.stDownloadButton button:hover, .stButton button:hover {
    background: var(--teal);
    color: #FFFFFF;
}

/* ---------- Dataframe ---------- */

[data-testid="stDataFrame"] {
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
}

</style>
""", unsafe_allow_html=True)

# ============================================================
# Hero
# ============================================================

st.markdown("""
<div class="hero">
    <div class="hero-eyebrow">Territory &middot; Cluster &middot; Field Visits</div>
    <h1>🧭 Field Visit Compliance</h1>
    <p>Upload the daily CRM export to check first-half and second-half field time against
    target, spot bunched check-ins, and export a formatted report for review. Optionally
    upload the credit (TCM submission) report to add login counts per CM.</p>
    <div class="route-line"></div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# Upload
# ============================================================

upload_col1, upload_col2 = st.columns(2)

with upload_col1:
    uploaded_file = st.file_uploader(
        "Upload CRM file (.xlsx or .csv)",
        type=["xlsx", "csv"],
        help="Expects columns: Territory, Cluster, EmployeeName, City, Lead Date, Lead Time (optionally DM Name)",
        key="crm_uploader"
    )

with upload_col2:
    credit_file = st.file_uploader(
        "Upload Credit Report (.xlsx or .csv) — optional",
        type=["xlsx", "csv"],
        help=(
            "Used to count logins. Looks for a 'Submitted to TCM Date Time' "
            "column (falls back to Excel column AL) and a 'CL Name' column "
            "(falls back to Excel column BE)."
        ),
        key="credit_uploader"
    )

# ============================================================
# Date criteria selection
# ============================================================

st.markdown('<div class="section-label" style="margin-top:0.6rem;">Date range</div>', unsafe_allow_html=True)

date_choice = st.radio(
    "Which visits should be included?",
    options=["Today", "Yesterday", "Custom range"],
    horizontal=True,
    label_visibility="collapsed",
    key="date_choice"
)

now_ist = pd.Timestamp.now(tz="Asia/Kolkata")
today_ist = now_ist.date()

if date_choice == "Today":
    date_from = today_ist
    date_to = today_ist
elif date_choice == "Yesterday":
    date_from = today_ist - pd.Timedelta(days=1)
    date_to = today_ist - pd.Timedelta(days=1)
else:
    range_col1, range_col2 = st.columns(2)
    with range_col1:
        date_from = st.date_input("From", value=today_ist, key="date_from_input")
    with range_col2:
        date_to = st.date_input("To", value=today_ist, key="date_to_input")

    if date_from > date_to:
        st.error("The 'From' date is after the 'To' date. Please fix the range.")
        st.stop()

# Calendar days covered by the selected range (for display only). The
# actual FH/SH hour requirement is based on *working* days, which can only
# be determined once the CRM file is loaded (a date counts as a holiday if
# nobody in the file logged a visit on it) — see the "Working Days Count"
# block further down.
num_days = (date_to - date_from).days + 1
HOURS_PER_DAY_PER_HALF = 2

# Distance threshold above which we flag instead of showing the number
DISTANCE_FLAG_KM = 200
DISTANCE_FLAG_LABEL = ">200 km"

# Rough bounding box for India, used to sanity-check that a lat/long pair
# is plausibly within the country. Coordinates outside this box (or
# missing/unparseable coordinates) are treated as an invalid location.
INDIA_LAT_MIN, INDIA_LAT_MAX = 6.5, 37.6
INDIA_LON_MIN, INDIA_LON_MAX = 68.0, 97.5

# Columns that get summed into each DM's subtotal row and the grand
# total row (see "DM-wise subtotal rows" block below).
DM_TOTAL_SUM_COLS = [
    "Logins",
    "SH Logins",
    "Total Logins",
]

# ------------------------------------------------------------
# Territory -> DM Name mapping.
# Neither the CRM export nor the credit report carries a DM Name column,
# so DM Name is derived from each row's Territory using this fixed table
# instead. Update this dict whenever territories are added/reassigned to
# a different DM. Matching is case/whitespace-insensitive; any Territory
# not listed here falls back to "Unassigned" and is flagged with a
# one-time warning so it's easy to spot a missing mapping.
# ------------------------------------------------------------
TERRITORY_TO_DM = {
    "khammam": "Anji",
    "suryapet": "Anji",
    "nalgonda": "Anji",
    "mahabubnagar": "Anji",
    "wanaparthy": "Anji",
    "nagarkurnool": "Anji",
    "jagityal": "Ravoof",
    "karimnagar": "Ravoof",
    "mancherial": "Ravoof",
    "nizamabad": "Ravoof",
    "warangal": "Ravoof",
    "narsampet": "Ravoof",
    "vijayawada": "Vijayawada",
}


def get_dm_name(territory):
    """Looks up the DM Name for a Territory via TERRITORY_TO_DM (case/whitespace
    -insensitive). Returns 'Unassigned' if the territory isn't in the table."""
    if pd.isna(territory):
        return "Unassigned"
    return TERRITORY_TO_DM.get(str(territory).strip().lower(), "Unassigned")

if date_from == date_to:
    st.caption(f"Showing visits for: {date_from.strftime('%d-%b-%Y')}")
else:
    st.caption(
        f"Showing visits from {date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')} "
        f"({num_days} calendar day(s)). Required FH/SH hours will be based on the "
        f"actual working-day count once the CRM file is uploaded — any date with "
        f"no login activity from anyone is treated as a holiday and excluded."
    )

if not uploaded_file:
    st.markdown("""
    <div class="card">
        <div class="section-label">Waiting for a file</div>
        <p style="color:var(--muted); margin:0;">
        Drop the CRM export above to generate the compliance summary for the
        selected date range. Field-half (FH) visits before 2 PM need at least
        2 hours in-field per working day; second-half (SH) visits need at
        least 2 hours per working day. A date with zero login activity from
        anyone is treated as a holiday and excluded from the working-day
        count, so the required FH/SH hours scale with the actual number of
        working days in the range — not just its calendar length. The
        credit report is optional — upload it too if you want login counts
        included.
        </p>
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# Helper: normalize LoginDate to IST, regardless of incoming format
# ============================================================

def normalize_to_ist(series: pd.Series) -> pd.Series:
    """
    Parses a datetime column that may arrive in mixed/inconsistent formats and
    returns a clean, timezone-naive datetime series expressed in IST.

    Rules:
    - Tries day-first parsing (DD-MM-YYYY style) first, since that's the
      expected default for this CRM export.
    - Any values that fail to parse are retried without the day-first
      assumption (covers stray MM-DD-YYYY / ISO / other rows).
    - If a value carries explicit timezone info (e.g. 'Z', '+00:00', a UTC
      offset), it is converted into IST.
    - If a value is already timezone-naive, it is assumed to already
      represent local/IST time and is left unchanged.
    """

    if series.empty:
        return series

    # First pass: day-first parsing
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)

    # Second pass: retry any still-unparsed (non-null original) values
    # without the day-first assumption
    mask = parsed.isna() & series.notna()
    if mask.any():
        retry = pd.to_datetime(series[mask], errors="coerce", dayfirst=False)
        parsed.loc[mask] = retry

    # If the whole column parsed as tz-aware (uniform tz across all rows),
    # convert straight to IST and drop the tz label
    if getattr(parsed.dt, "tz", None) is not None:
        parsed = parsed.dt.tz_convert("Asia/Kolkata").dt.tz_localize(None)
        return parsed

    # Otherwise, values may be a mix of naive and tz-aware datetimes
    # (this can happen when the source column has mixed formats).
    # Handle element-wise as a fallback.
    def _fix(x):
        if pd.isna(x):
            return x
        tzinfo = getattr(x, "tzinfo", None)
        if tzinfo is not None:
            return x.tz_convert("Asia/Kolkata").tz_localize(None)
        return x

    parsed = parsed.apply(_fix)
    return parsed


# ============================================================
# Helper: read a CSV or Excel upload with encoding/delimiter fallbacks
# ============================================================

def read_uploaded_table(file, label):
    """
    Reads a CSV or Excel file-like upload into a DataFrame, trying a
    sequence of encodings/delimiters for CSVs. Returns the DataFrame,
    or (None, error) on failure.
    """
    is_csv = file.name.lower().endswith(".csv")

    if is_csv:
        read_attempts = [
            {"encoding": "utf-8-sig", "sep": None, "engine": "python"},
            {"encoding": "utf-8", "sep": None, "engine": "python"},
            {"encoding": "latin-1", "sep": None, "engine": "python"},
            {"encoding": "cp1252", "sep": None, "engine": "python"},
        ]

        df_local = None
        last_error = None

        for attempt in read_attempts:
            try:
                file.seek(0)
                df_local = pd.read_csv(file, **attempt)
                break
            except Exception as e:
                last_error = e
                continue

        if df_local is None:
            st.error(
                f"Couldn't read the {label} CSV file. It may use an unusual "
                f"encoding or delimiter. Please re-export it as UTF-8 CSV, "
                f"or upload the original Excel (.xlsx) file instead."
            )
            st.exception(last_error)
            return None
        return df_local
    else:
        try:
            file.seek(0)
            return pd.read_excel(file)
        except Exception as e:
            st.error(f"Couldn't read the {label} Excel file. Please check the file and try again.")
            st.exception(e)
            return None


# ============================================================
# Helper: locate a column by name, falling back to an Excel column letter
# ============================================================

def _col_letter_to_index(letter):
    """Convert an Excel column letter (e.g. 'AL') to a zero-based index."""
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def find_column(df, preferred_name, fallback_letter):
    """
    Finds a column by (case/whitespace-insensitive) name match first;
    if not found, falls back to the column at the given Excel letter
    position. Returns the actual column label, or None if unavailable.
    """
    normalized = {str(c).strip().lower(): c for c in df.columns}
    key = preferred_name.strip().lower()
    if key in normalized:
        return normalized[key]

    idx = _col_letter_to_index(fallback_letter)
    if 0 <= idx < len(df.columns):
        return df.columns[idx]

    return None


def find_column_by_name(df, preferred_name):
    """
    Finds a column by (case/whitespace-insensitive) name match only —
    no Excel-letter fallback. Returns the actual column label, or None
    if no column with that name exists.
    """
    normalized = {str(c).strip().lower(): c for c in df.columns}
    return normalized.get(preferred_name.strip().lower())


# ============================================================
# Helper: total in-field minutes, summed per calendar day
# ============================================================

def daily_summed_minutes(session_df):
    """
    Given a subset of rows for one CM/one session (FH or SH), computes
    (last visit - first visit) *per calendar day* and sums those daily
    spans together. This is what gets compared against the multi-day
    target (num_days x 2 hours), instead of naively taking the first and
    last timestamp across the whole selected range (which would overstate
    "in-field time" whenever the range covers more than one day).
    """
    if session_df.empty:
        return 0
    per_day = session_df.groupby(session_df["LoginDate"].dt.date)["LoginDate"].agg(
        lambda s: (s.max() - s.min()).total_seconds() / 60
    )
    return per_day.sum()


# ============================================================
# Helper: count of working days on which a CM is compliant
# ============================================================

def daily_compliance_details(fh_df, sh_df, working_dates):
    """
    Calculates compliance day counts once per CM.

    A day is compliant only when both FH and SH meet the 2-hour target.
    The same result is used by Summary and Location Summary so their
    compliant/non-compliant counts always match.
    """
    threshold = HOURS_PER_DAY_PER_HALF * 60

    fh_daily = (
        fh_df.groupby(fh_df["LoginDate"].dt.date)["LoginDate"]
        .agg(lambda s: (s.max() - s.min()).total_seconds() / 60)
        if not fh_df.empty else pd.Series(dtype=float)
    )

    sh_daily = (
        sh_df.groupby(sh_df["LoginDate"].dt.date)["LoginDate"]
        .agg(lambda s: (s.max() - s.min()).total_seconds() / 60)
        if not sh_df.empty else pd.Series(dtype=float)
    )

    compliant_days = 0
    non_compliant_days = 0

    for d in working_dates:
        fh_minutes = fh_daily.get(d, 0)
        sh_minutes = sh_daily.get(d, 0)

        if fh_minutes >= threshold and sh_minutes >= threshold:
            # Fully Compliant
            compliant_days += 1

        elif fh_minutes < threshold and sh_minutes < threshold:
            # Non-Compliant
            non_compliant_days += 1

        else:
            # Partially Compliant
            # Do not count as Non-Compliant
            pass

    return compliant_days, non_compliant_days


# ============================================================
# Helper: great-circle distance between two lat/long points, in km
# ============================================================

def haversine_km(lat1, lon1, lat2, lon2):
    """
    Standard haversine great-circle distance between two points given in
    decimal degrees. Returns kilometers. Any NaN input returns 0 so a
    missing ping doesn't blow up a whole day's total.
    """
    if pd.isna(lat1) or pd.isna(lon1) or pd.isna(lat2) or pd.isna(lon2):
        return 0.0

    R = 6371.0  # Earth radius, km
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)

    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


# ============================================================
# Helper: is a lat/long pair a plausible, in-country location?
# ============================================================

def is_valid_india_coord(lat, lon):
    """
    Returns False if lat/lon is missing/unparseable, or falls outside a
    rough bounding box for India. Used to flag pings that are clearly bad
    data (0,0 / null island, GPS glitches) or genuinely outside the
    country, so the distance calc doesn't silently produce a misleading
    number for them.
    """
    if pd.isna(lat) or pd.isna(lon):
        return False
    if not (INDIA_LAT_MIN <= lat <= INDIA_LAT_MAX):
        return False
    if not (INDIA_LON_MIN <= lon <= INDIA_LON_MAX):
        return False
    return True


# ============================================================
# Helper: distance between the first visit and the last visit of a session
# ============================================================

def first_last_distance(session_df, lat_col, lon_col):
    """
    Given a subset of rows for one CM/one session (FH or SH), computes the
    straight-line (haversine) distance between the row with the earliest
    LoginDate (the session's "first visit") and the row with the latest
    LoginDate (the "last visit") — mirroring the First Visit / Last Visit
    columns already shown in the summary, rather than summing distance
    across every intermediate ping.

    Returns (distance_km, is_invalid):
    - distance_km is 0.0 whenever there's nothing to compare (0-1 visits)
      or either endpoint's coordinates are invalid.
    - is_invalid is True when there are 2+ visits but either the first or
      last visit's coordinates are missing or fall outside India's
      bounding box — meaning the distance figure can't be trusted.
    """
    if session_df.empty or lat_col is None or lon_col is None or len(session_df) < 2:
        return 0.0, False

    first_row = session_df.loc[session_df["LoginDate"].idxmin()]
    last_row = session_df.loc[session_df["LoginDate"].idxmax()]

    lat1, lon1 = first_row[lat_col], first_row[lon_col]
    lat2, lon2 = last_row[lat_col], last_row[lon_col]

    if not is_valid_india_coord(lat1, lon1) or not is_valid_india_coord(lat2, lon2):
        return 0.0, True

    return haversine_km(lat1, lon1, lat2, lon2), False


def format_distance(km_value):
    """
    Returns the display value for a distance-covered figure: the flagged
    label (">200 km") when it exceeds the threshold, otherwise the value
    rounded to 2 decimal places. Keeping the flagged case as a distinct
    string (rather than the number) lets both the on-screen preview and
    the exported Excel highlight it in red.
    """
    if km_value > DISTANCE_FLAG_KM:
        return DISTANCE_FLAG_LABEL
    return round(km_value, 2)


# ============================================================
# Build login counts (FH / SH) per CL Name from the credit report,
# restricted to the CRM report's own date. The actual processing
# happens below, once we know which day the CRM export covers.
# ============================================================

login_lookup = {}
credit_warning = None


def build_login_lookup(credit_file, date_from, date_to):
    """
    Reads the credit report, keeps only rows whose 'Submitted to TCM
    Date Time' falls within [date_from, date_to] (inclusive, IST), and
    returns (login_lookup, warning). login_lookup maps a normalized
    CL Name -> {"FH": n, "SH": n, "DisplayName": ..., "Territory": ...,
    "Cluster": ..., "DM Name": ...}. Territory/Cluster/DisplayName/DM
    Name are carried along so that employees present only in the credit
    report (no matching CM in the CRM file) can still be listed in the
    summary.
    """
    lookup = {}
    warning = None

    credit_df = read_uploaded_table(credit_file, "credit report")
    if credit_df is None:
        return lookup, warning

    credit_df.columns = [str(c).strip() for c in credit_df.columns]

    submitted_col = find_column(credit_df, "Submitted to TCM Date Time", "AL")
    name_col = find_column(credit_df, "CL Name", "BE")
    territory_col = find_column_by_name(credit_df, "Territory")
    cluster_col = find_column_by_name(credit_df, "Cluster Name")

    if submitted_col is None or name_col is None:
        warning = (
            "Couldn't locate the login timestamp and/or CL Name columns "
            "in the credit report (expected 'Submitted to TCM Date Time' "
            "around column AL, and 'CL Name' around column BE). Login "
            "counts will show as 0."
        )
        return lookup, warning

    if territory_col is None or cluster_col is None:
        missing = []
        if territory_col is None:
            missing.append("'Territory'")
        if cluster_col is None:
            missing.append("'Cluster Name'")
        warning = (
            f"Couldn't locate {' and '.join(missing)} column(s) in the "
            f"credit report. Any employee added from the credit report "
            f"only (no matching CM in the CRM file) will show a blank "
            f"value for that field."
        )

    credit_df[submitted_col] = normalize_to_ist(credit_df[submitted_col])
    credit_df = credit_df.dropna(subset=[submitted_col])

    # Keep only logins submitted within the selected date range
    credit_dates = credit_df[submitted_col].dt.date
    credit_df = credit_df[(credit_dates >= date_from) & (credit_dates <= date_to)]

    if credit_df.empty:
        if date_from == date_to:
            range_label = date_from.strftime('%d-%b-%Y')
        else:
            range_label = f"{date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')}"
        warning = (
            f"No rows in the credit report matched the selected date range "
            f"({range_label}). Login counts will show as 0. "
            f"Double-check the credit report covers this period."
        )
        return lookup, warning

    credit_df["_Session"] = credit_df[submitted_col].dt.hour.apply(
        lambda x: "FH" if x < 14 else "SH"
    )
    credit_df["_NameKey"] = (
        credit_df[name_col].astype(str).str.strip().str.lower()
    )
    credit_df["_DisplayName"] = credit_df[name_col].astype(str).str.strip()
    credit_df["_Territory"] = (
        credit_df[territory_col].astype(str).str.strip() if territory_col else ""
    )
    credit_df["_Cluster"] = (
        credit_df[cluster_col].astype(str).str.strip() if cluster_col else ""
    )
    credit_df["_DMName"] = (
        credit_df["_Territory"].apply(get_dm_name) if territory_col else "Unassigned"
    )

    counts = (
        credit_df.groupby(["_NameKey", "_Session"])
        .size()
        .unstack(fill_value=0)
    )
    meta = credit_df.groupby("_NameKey").agg(
        {"_DisplayName": "first", "_Territory": "first", "_Cluster": "first", "_DMName": "first"}
    )

    for name_key in counts.index:
        row = counts.loc[name_key]
        m = meta.loc[name_key]
        lookup[name_key] = {
            "FH": int(row.get("FH", 0)),
            "SH": int(row.get("SH", 0)),
            "DisplayName": m["_DisplayName"],
            "Territory": m["_Territory"],
            "Cluster": m["_Cluster"],
            "DM Name": m["_DMName"],
        }

    return lookup, warning


if uploaded_file:

    df = read_uploaded_table(uploaded_file, "CRM")
    if df is None:
        st.stop()

    # Normalize column names (strips stray whitespace/BOM artifacts that
    # sometimes survive in CSV headers)
    df.columns = [str(c).strip() for c in df.columns]

    required_cols = {"Territory", "Cluster", "EmployeeName", "City", "Lead Date", "Lead Time"}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        st.error(
            f"This file is missing required column(s): {', '.join(sorted(missing_cols))}. "
            f"Found columns: {', '.join(df.columns)}"
        )
        st.stop()

    # DM Name isn't present in the CRM export, so it's derived from each
    # row's Territory via TERRITORY_TO_DM instead. The Summary sheet is
    # grouped by the resulting DM with a subtotal row after each DM's
    # rows, plus a Grand Total row at the end.
    has_dm = True
    df["DM Name"] = df["Territory"].apply(get_dm_name)

    unmapped_territories = sorted(
        t for t in df.loc[df["DM Name"] == "Unassigned", "Territory"].astype(str).str.strip().unique()
        if t
    )
    if unmapped_territories:
        st.warning(
            f"No DM mapping found for territory/territories: {', '.join(unmapped_territories)}. "
            f"These rows are grouped under 'Unassigned' — add them to TERRITORY_TO_DM in the "
            f"script to assign the correct DM."
        )

    # "Lead Date" carries the actual visit date; "Lead Time" is a separate
    # submission timestamp whose date part is just whenever the row was
    # exported/synced (often the same day for the whole file) — only its
    # time-of-day is meaningful. The rest of the app refers to the combined,
    # correct visit timestamp internally as "LoginDate": Lead Date's date +
    # Lead Time's time-of-day.
    #
    # "Lead Time" can arrive in two different shapes depending on the
    # source file: as a pure time-of-day value (pandas/openpyxl read this
    # as a timedelta64 when Excel stores it as a time-only cell), or as a
    # full datetime/text timestamp (handled by normalize_to_ist, as
    # before). Treating a timedelta column as if it were a datetime (the
    # old behavior) makes pd.to_datetime silently return NaT for every
    # row, which then wipes out the entire file once LoginDate rows with
    # NaT are dropped further down — so detect which shape we have and
    # extract the time-of-day component accordingly.
    _lead_date = normalize_to_ist(df["Lead Date"])

    if pd.api.types.is_timedelta64_dtype(df["Lead Time"]):
        _lead_time_of_day = df["Lead Time"]
    else:
        _lead_time = normalize_to_ist(df["Lead Time"])
        _lead_time_of_day = _lead_time - _lead_time.dt.normalize()

    df["LoginDate"] = _lead_date.dt.normalize() + _lead_time_of_day

    # ------------------------------------------------------------
    # Locate latitude/longitude columns for distance-covered calc.
    # Named-header match first ("Latitude" / "Longitude"), falling back
    # to Excel columns BG / BH if no matching header is found. This is
    # optional — if neither can be located, distance columns show as 0
    # and a one-time warning is displayed further down.
    # ------------------------------------------------------------
    lat_col = find_column(df, "Latitude", "BG")
    lon_col = find_column(df, "Longitude", "BH")

    has_geo = lat_col is not None and lon_col is not None
    geo_warning = None

    if has_geo:
        df[lat_col] = pd.to_numeric(df[lat_col], errors="coerce")
        df[lon_col] = pd.to_numeric(df[lon_col], errors="coerce")
        # If every value failed to parse as numeric, treat as not available
        if df[lat_col].notna().sum() == 0 or df[lon_col].notna().sum() == 0:
            has_geo = False
            geo_warning = (
                "Found latitude/longitude columns, but couldn't parse any "
                "values as numbers. Distance Covered will show as 0."
            )
    else:
        geo_warning = (
            "Couldn't locate Latitude/Longitude columns in the CRM file "
            "(expected named columns, or around Excel columns BG/BH). "
            "Distance Covered will show as 0."
        )

    # ------------------------------------------------------------
    # Locate a Loan Requirement - Yes/No column, used to count "Leads" in
    # the Location Summary sheet (Custom range only). Name match only —
    # no letter fallback, since this column's position isn't fixed for
    # this export. A row is counted as a lead if the value == "Yes"
    # (case/whitespace-insensitive).
    # ------------------------------------------------------------
    loan_requirement_col = find_column_by_name(df, "Loan Requirement - Yes/No")

    # Remove rows where Lead Date or Lead Time failed to parse (LoginDate is NaT)
    df = df.dropna(subset=["LoginDate"])

    # Keep only CRM rows that fall within the selected date range
    df = df[
        (df["LoginDate"].dt.date >= date_from) & (df["LoginDate"].dt.date <= date_to)
    ]

    if df.empty:
        if date_from == date_to:
            range_label = date_from.strftime('%d-%b-%Y')
        else:
            range_label = f"{date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')}"
        st.warning(
            f"No CRM rows fall within the selected date range ({range_label}). "
            f"Try a different range."
        )
        st.stop()

    # ------------------------------------------------------------
    # Actual working days within the selected range: any date on which
    # *no employee at all* has a login is treated as a holiday (Sundays,
    # off-Saturdays, etc.) and excluded from the working-day count. The
    # FH/SH requirement for every CM is then working_days_count x 2 hrs.
    # ------------------------------------------------------------
    all_range_dates = list(pd.date_range(date_from, date_to).date)
    working_dates = sorted(df["LoginDate"].dt.date.unique())
    holiday_dates = [d for d in all_range_dates if d not in working_dates]

    working_days_count = len(working_dates)
    fh_required_minutes = working_days_count * HOURS_PER_DAY_PER_HALF * 60
    sh_required_minutes = working_days_count * HOURS_PER_DAY_PER_HALF * 60

    required_msg = (
        f"Working Days Count: {working_days_count} of {num_days} day(s) in range. "
        f"Required per CM: FH {fh_required_minutes // 60}:{fh_required_minutes % 60:02d} hrs, "
        f"SH {sh_required_minutes // 60}:{sh_required_minutes % 60:02d} hrs."
    )

    if holiday_dates:
        holiday_label = ", ".join(d.strftime("%d-%b") for d in holiday_dates)
        st.info(f"{required_msg} No login activity found on: {holiday_label} (treated as holiday).")
    else:
        st.caption(f"{required_msg} No holidays detected in this range.")

    if geo_warning:
        st.warning(geo_warning)

    if date_choice == "Custom range" and loan_requirement_col is None:
        st.warning(
            "Couldn't locate a 'Loan Requirement - Yes/No' column in the CRM file, "
            "so the Leads count in the Location Summary sheet will show as 0."
        )

    # Only count logins from the credit report that fall within the same
    # selected date range — not whichever date happens to dominate the file.
    if credit_file is not None:
        login_lookup, credit_warning = build_login_lookup(credit_file, date_from, date_to)
        if date_from == date_to:
            login_range_label = date_from.strftime('%d-%b-%Y')
        else:
            login_range_label = f"{date_from.strftime('%d-%b-%Y')} to {date_to.strftime('%d-%b-%Y')}"
        st.caption(f"Login counts are restricted to: {login_range_label}")

    if credit_warning:
        st.warning(credit_warning)

    # FH before 2 PM
    df["Session"] = df["LoginDate"].dt.hour.apply(
        lambda x: "FH" if x < 14 else "SH"
    )

    summary = []
    matched_login_keys = set()

    # Per-CM lookup keyed by (Territory, Cluster, lowercased EmployeeName),
    # populated in the loop below and consumed later when building the
    # Location Summary sheet.
    daily_compliance_lookup = {}

    group_cols = (
        ["DM Name", "Territory", "Cluster", "EmployeeName"]
        if has_dm else
        ["Territory", "Cluster", "EmployeeName"]
    )

    for key, grp in df.groupby(group_cols):

        if has_dm:
            dm_name, territory, cluster, emp = key
        else:
            dm_name = ""
            territory, cluster, emp = key

        fh = grp[grp["Session"] == "FH"]
        sh = grp[grp["Session"] == "SH"]

        # Working-day compliance count for this CM, used by the Location
        # Summary sheet (Custom range only).
        cm_key = (territory, cluster, str(emp).strip().lower())
        compliant_days, non_compliant_days = daily_compliance_details(
            fh, sh, working_dates
        )
        daily_compliance_lookup[cm_key] = {
            "Compliant Days": compliant_days,
            "Non-Compliant Days": non_compliant_days,
        }

        # ---------------- FH ----------------

        fh_locations = ", ".join(
            fh["City"].dropna().astype(str).unique()
        )

        fh_visits = len(fh)

        if len(fh):
            fh_first = fh["LoginDate"].min()
            fh_last = fh["LoginDate"].max()

            # Sum of each day's own (last - first) span, not the span of
            # the whole selected range — see daily_summed_minutes().
            fh_minutes = int(daily_summed_minutes(fh))

            fh_duration = f"{fh_minutes // 60}:{fh_minutes % 60:02d}"
            fh_status = "Met" if fh_minutes >= fh_required_minutes else "Not Met"
        else:
            fh_first = pd.NaT
            fh_last = pd.NaT
            fh_minutes = 0
            fh_duration = "00:00"
            fh_status = "Not Met"

        # Distance is the straight-line distance between the FH first
        # visit and the FH last visit (not a sum across every ping) —
        # see first_last_distance(). If either endpoint's coordinates
        # are missing or fall outside India, the location is flagged
        # invalid instead of producing a distance number.
        if has_geo:
            fh_distance_km, fh_invalid = first_last_distance(fh, lat_col, lon_col)
        else:
            fh_distance_km, fh_invalid = 0.0, False

        fh_distance_display = "" if fh_invalid else format_distance(fh_distance_km)

        # ---------------- SH ----------------

        sh_locations = ", ".join(
            sh["City"].dropna().astype(str).unique()
        )

        sh_visits = len(sh)

        if len(sh):
            sh_first = sh["LoginDate"].min()
            sh_last = sh["LoginDate"].max()

            sh_minutes = int(daily_summed_minutes(sh))

            sh_duration = f"{sh_minutes // 60}:{sh_minutes % 60:02d}"
            sh_status = "Met" if sh_minutes >= sh_required_minutes else "Not Met"
        else:
            sh_first = pd.NaT
            sh_last = pd.NaT
            sh_minutes = 0
            sh_duration = "00:00"
            sh_status = "Not Met"

        if has_geo:
            sh_distance_km, sh_invalid = first_last_distance(sh, lat_col, lon_col)
        else:
            sh_distance_km, sh_invalid = 0.0, False

        sh_distance_display = "" if sh_invalid else format_distance(sh_distance_km)

        # ---------------- Total ----------------

        total = fh_visits + sh_visits

        # Distinct calendar dates this CM has any login activity on
        # (FH or SH combined) within the selected range.
        employee_working_days = grp["LoginDate"].dt.date.nunique()

        # ---------------- Logins (from credit report) ----------------

        name_key = str(emp).strip().lower()
        login_counts = login_lookup.get(name_key, {"FH": 0, "SH": 0})
        if name_key in login_lookup:
            matched_login_keys.add(name_key)

        fh_logins = login_counts["FH"]
        sh_logins = login_counts["SH"]
        total_logins = fh_logins + sh_logins

        if fh_status == "Met" and sh_status == "Met":
            compliance = "Fully Compliant"
        elif fh_status == "Met" or sh_status == "Met":
            compliance = "Partially Compliant"
        else:
            compliance = "Non-Compliant"

        total_non_compliances = daily_compliance_lookup.get(
            cm_key,
            {"Non-Compliant Days": working_days_count}
        )["Non-Compliant Days"]

        # ---------------- Remarks ----------------

        remarks_list = []

        if len(fh) > 1:
            fh_sorted = fh.sort_values("LoginDate")
            fh_gaps = (
                fh_sorted["LoginDate"]
                .diff()
                .dt.total_seconds()
                .dropna() / 60
            )
            if fh_gaps.mean() < 5:
                remarks_list.append(
                    "FH visits bunched (avg gap <5m); verify field presence."
                )

        if len(sh) > 1:
            sh_sorted = sh.sort_values("LoginDate")
            sh_gaps = (
                sh_sorted["LoginDate"]
                .diff()
                .dt.total_seconds()
                .dropna() / 60
            )
            if sh_gaps.mean() < 5:
                remarks_list.append(
                    "SH visits bunched (avg gap <5m); verify field presence."
                )

        if fh_invalid:
            remarks_list.append("FH: Invalid Location.")

        if sh_invalid:
            remarks_list.append("SH: Invalid Location.")

        remarks = " ".join(remarks_list)

        summary.append([
            dm_name,
            territory,
            cluster,
            emp,
            fh_locations,
            fh_visits,
            fh_first,
            fh_last,
            fh_duration,
            fh_status,
            fh_logins,
            fh_distance_display,
            sh_locations,
            sh_visits,
            sh_first,
            sh_last,
            sh_duration,
            sh_status,
            sh_logins,
            sh_distance_display,
            total,
            total_logins,
            employee_working_days,
            working_days_count,
            compliance,
            total_non_compliances,
            remarks
        ])

    # ------------------------------------------------------------
    # Employees who appear in the credit report but have no matching CM
    # in the CRM file are still included, using Territory/Cluster/Name
    # from the credit report. Their CRM-based stats (visits, durations,
    # working days, distance) are 0/blank since there's no CRM data to
    # derive them from; login counts still come through from the credit
    # report.
    # ------------------------------------------------------------
    unmatched_keys = sorted(set(login_lookup.keys()) - matched_login_keys)

    for name_key in unmatched_keys:
        info = login_lookup[name_key]

        fh_logins = info["FH"]
        sh_logins = info["SH"]
        total_logins = fh_logins + sh_logins

        fh_status = "Met" if 0 >= fh_required_minutes else "Not Met"
        sh_status = "Met" if 0 >= sh_required_minutes else "Not Met"

        if fh_status == "Met" and sh_status == "Met":
            compliance = "Fully Compliant"
        elif fh_status == "Met" or sh_status == "Met":
            compliance = "Partially Compliant"
        else:
            compliance = "Non-Compliant"

        summary.append([
            (info.get("DM Name", "") or "") if has_dm else "",
            info.get("Territory", "") or "",
            info.get("Cluster", "") or "",
            info.get("DisplayName", name_key),
            "",       # FH Locations Visited
            0,        # FH Visits
            pd.NaT,   # FH First Visit
            pd.NaT,   # FH Last Visit
            "00:00",  # FH Duration
            fh_status,
            fh_logins,
            0.0,      # FH Distance (km)
            "",       # SH Locations Visited
            0,        # SH Visits
            pd.NaT,   # SH First Visit
            pd.NaT,   # SH Last Visit
            "00:00",  # SH Duration
            sh_status,
            sh_logins,
            0.0,      # SH Distance (km)
            0,        # Total Visits (Day)
            total_logins,
            0,        # Employee Working Days (no CRM activity)
            working_days_count,
            compliance,
            working_days_count,
            "No CRM visit data for this employee — added from Credit Report."
        ])

    cols = [
        "DM Name",
        "Territory",
        "Cluster",
        "CM Name",
        "FH Locations Visited",
        "FH Visits",
        "FH First Visit",
        "FH Last Visit",
        "FH Duration",
        "FH Status",
        "Logins",
        "FH Distance (km)",
        "SH Locations Visited",
        "SH Visits",
        "SH First Visit",
        "SH Last Visit",
        "SH Duration",
        "SH Status",
        "SH Logins",
        "SH Distance (km)",
        "Total Visits (Day)",
        "Total Logins",
        "Employee Working Days",
        "Working Days Count",
        "Overall Compliance",
        "Total Non-Compliances",
        "Remarks"
    ]

    summary_df = pd.DataFrame(summary, columns=cols)

    if not has_dm:
        # No DM Name column found in the source file — drop it rather
        # than showing an empty column.
        summary_df = summary_df.drop(columns=["DM Name"])
        sort_cols = ["Territory", "Cluster", "CM Name"]
    else:
        sort_cols = ["DM Name", "Territory", "Cluster", "CM Name"]

    summary_df = summary_df.sort_values(sort_cols).reset_index(drop=True)

    # Show first/last visit as time-only for a single-day range, but
    # include the date too when the range spans multiple days (since a
    # bare "HH:MM" would be ambiguous about which day it refers to).
    time_cols = [
        "FH First Visit",
        "FH Last Visit",
        "SH First Visit",
        "SH Last Visit"
    ]

    time_format = "%H:%M" if date_from == date_to else "%d-%b %H:%M"

    for col in time_cols:
        summary_df[col] = pd.to_datetime(
            summary_df[col],
            errors="coerce"
        ).dt.strftime(time_format).fillna("")

    # ------------------------------------------------------------
    # DM-wise subtotal rows + Grand Total row (Summary sheet only).
    # Only added when a "DM Name" column was found in the CRM file.
    # A subtotal row is inserted after each DM's block of CM rows,
    # labeled "DM-1 Total", "DM-2 Total", ... in the S.No column, with
    # the numeric columns in DM_TOTAL_SUM_COLS summed for that DM.
    # A final "Grand Total" row sums those same columns across every
    # DM. All other columns are left blank on total rows.
    # ------------------------------------------------------------
    if has_dm:
        blank_row = {c: "" for c in summary_df.columns}

        final_rows = []
        grand_totals = {c: 0 for c in DM_TOTAL_SUM_COLS}
        dm_index = 0

        for dm_name, dm_group in summary_df.groupby("DM Name", sort=False):
            for _, r in dm_group.iterrows():
                final_rows.append(r.to_dict())

            dm_index += 1
            total_row = dict(blank_row)
            total_row["S.No"] = f"DM-{dm_index} Total"
            for c in DM_TOTAL_SUM_COLS:
                dm_sum = int(pd.to_numeric(dm_group[c], errors="coerce").fillna(0).sum())
                total_row[c] = dm_sum
                grand_totals[c] += dm_sum
            final_rows.append(total_row)

        grand_total_row = dict(blank_row)
        grand_total_row["S.No"] = "Grand Total"
        for c in DM_TOTAL_SUM_COLS:
            grand_total_row[c] = grand_totals[c]
        final_rows.append(grand_total_row)

        summary_df = pd.DataFrame(final_rows, columns=["S.No"] + list(summary_df.columns))

        # Renumber only the real CM rows sequentially; total rows keep
        # their text label.
        running_no = 0
        new_sno = []
        for val in summary_df["S.No"]:
            if isinstance(val, str) and val.endswith("Total"):
                new_sno.append(val)
            else:
                running_no += 1
                new_sno.append(running_no)
        summary_df["S.No"] = new_sno
    else:
        summary_df.insert(0, "S.No", range(1, len(summary_df) + 1))

    # ------------------------------------------------------------
    # Location Summary sheet — Custom range only.
    # One row per CM/location. Compliance and non-compliance counts come
    # from the same daily calculation used by Summary. DM Name is carried
    # along from the Summary sheet and DM-wise subtotal + Grand Total
    # rows are appended, mirroring the Summary sheet's layout.
    # ------------------------------------------------------------
    location_df = None
    location_merge_sizes = []

    if date_choice == "Custom range":
        if loan_requirement_col is not None:
            df["_IsLead"] = (
                df[loan_requirement_col]
                .astype(str)
                .str.strip()
                .str.lower()
                .eq("yes")
            )
        else:
            df["_IsLead"] = False

        cm_cols = (
            ["DM Name", "Territory", "Cluster", "CM Name"]
            if has_dm else
            ["Territory", "Cluster", "CM Name"]
        )
        cm_order = summary_df[
            summary_df["S.No"].apply(lambda v: not (isinstance(v, str) and v.endswith("Total")))
        ][cm_cols].drop_duplicates()
        location_rows = []
        s_no = 0

        for _, cm_row in cm_order.iterrows():
            dm_name = cm_row["DM Name"] if has_dm else ""
            territory = cm_row["Territory"]
            cluster = cm_row["Cluster"]
            cm_name = cm_row["CM Name"]
            cm_key = (territory, cluster, str(cm_name).strip().lower())

            compliance_data = daily_compliance_lookup.get(
                cm_key,
                {
                    "Compliant Days": 0,
                    "Non-Compliant Days": working_days_count,
                },
            )

            total_non_compliances = compliance_data["Non-Compliant Days"]

            cm_df = df[
                (df["Territory"] == territory)
                & (df["Cluster"] == cluster)
                & (df["EmployeeName"] == cm_name)
            ].copy()

            fh = cm_df[cm_df["Session"] == "FH"]
            sh = cm_df[cm_df["Session"] == "SH"]

            login_counts = login_lookup.get(
                str(cm_name).strip().lower(),
                {"FH": 0, "SH": 0},
            )
            fh_logins = int(login_counts.get("FH", 0))
            sh_logins = int(login_counts.get("SH", 0))
            total_logins = fh_logins + sh_logins

            # Location-wise visit counts.
            fh_location_counts = (
                fh.groupby("City").size().to_dict()
                if not fh.empty else {}
            )
            sh_location_counts = (
                sh.groupby("City").size().to_dict()
                if not sh.empty else {}
            )

            # Lead count by location: Loan Requirement - Yes/No == Yes.
            lead_counts = (
                cm_df[cm_df["_IsLead"]]
                .groupby("City")
                .size()
                .to_dict()
                if not cm_df.empty else {}
            )

            locations = sorted(
                set(fh_location_counts.keys())
                | set(sh_location_counts.keys()),
                key=lambda x: str(x),
            )

            s_no += 1

            if not locations:
                row = [s_no]
                if has_dm:
                    row.append(dm_name)
                row += [territory, cluster, cm_name, "", 0, 0, total_logins, total_non_compliances]
                location_rows.append(row)
                location_merge_sizes.append(1)
            else:
                for location in locations:
                    fh_visits = int(fh_location_counts.get(location, 0))
                    sh_visits = int(sh_location_counts.get(location, 0))
                    leads = int(lead_counts.get(location, 0))
                    total_visits_loc = fh_visits + sh_visits

                    row = [s_no]
                    if has_dm:
                        row.append(dm_name)
                    row += [
                        territory,
                        cluster,
                        cm_name,
                        location,
                        total_visits_loc,
                        leads,
                        total_logins,
                        total_non_compliances,
                    ]
                    location_rows.append(row)

                location_merge_sizes.append(len(locations))

        loc_cols = ["S.No"]
        if has_dm:
            loc_cols.append("DM Name")
        loc_cols += [
            "Territory",
            "Cluster",
            "CM Name",
            "Locations Visited",
            "Total Visits",
            "Total Leads",
            "Total Logins",
            "Total Non-Compliances",
        ]

        location_df = pd.DataFrame(location_rows, columns=loc_cols)

        # ------------------------------------------------------------
        # DM-wise subtotal rows + Grand Total row, mirroring the Summary
        # sheet. "Total Visits" / "Total Leads" are per-location figures,
        # so they're summed straight across every location row for the
        # DM. "Total Logins" / "Total Non-Compliances" are per-CM
        # figures that repeat across a CM's location rows, so those are
        # summed once per unique CM instead of once per row (to avoid
        # double-counting a CM with multiple locations).
        # ------------------------------------------------------------
        if has_dm and not location_df.empty:
            blank_row = {c: "" for c in location_df.columns}
            final_rows = []
            new_merge_sizes = []
            grand_totals = {
                "Total Visits": 0,
                "Total Leads": 0,
                "Total Logins": 0,
                "Total Non-Compliances": 0,
            }
            dm_index = 0

            for dm_name_val, dm_group in location_df.groupby("DM Name", sort=False):
                for _, cm_group in dm_group.groupby("S.No", sort=False):
                    for _, r in cm_group.iterrows():
                        final_rows.append(r.to_dict())
                    new_merge_sizes.append(len(cm_group))

                dm_index += 1
                total_row = dict(blank_row)
                total_row["S.No"] = f"DM-{dm_index} Total"
                total_row["DM Name"] = dm_name_val
                total_row["Total Visits"] = int(dm_group["Total Visits"].sum())
                total_row["Total Leads"] = int(dm_group["Total Leads"].sum())

                cm_unique = dm_group.drop_duplicates(subset=["Territory", "Cluster", "CM Name"])
                total_row["Total Logins"] = int(cm_unique["Total Logins"].sum())
                total_row["Total Non-Compliances"] = int(cm_unique["Total Non-Compliances"].sum())

                for k in grand_totals:
                    grand_totals[k] += total_row[k]

                final_rows.append(total_row)
                new_merge_sizes.append(1)

            grand_total_row = dict(blank_row)
            grand_total_row["S.No"] = "Grand Total"
            for k, v in grand_totals.items():
                grand_total_row[k] = v
            final_rows.append(grand_total_row)
            new_merge_sizes.append(1)

            location_df = pd.DataFrame(final_rows, columns=loc_cols)
            location_merge_sizes = new_merge_sizes

    # Let the user know if some credit-report logins had no matching CM in
    # the CRM file — they've been added as extra rows above/below using
    # their credit-report details (Territory/Cluster/Name), with CRM-based
    # stats left at 0.
    if credit_file is not None and unmatched_keys:
        st.info(
            f"{len(unmatched_keys)} name(s) in the credit report didn't match "
            f"any CM in the CRM file. They've been added to the summary using "
            f"their Territory/Cluster/Name from the credit report — CRM-based "
            f"fields (visits, durations, working days, distance) show as 0 "
            f"for them."
        )

    # ============================================================
    # KPI strip
    # ============================================================

    is_total_row_mask = summary_df["S.No"].apply(lambda v: isinstance(v, str) and v.endswith("Total"))
    cm_rows_df = summary_df[~is_total_row_mask]

    total_cms = len(cm_rows_df)
    full_count = int((cm_rows_df["Overall Compliance"] == "Fully Compliant").sum())
    partial_count = int((cm_rows_df["Overall Compliance"] == "Partially Compliant").sum())
    non_count = int((cm_rows_df["Overall Compliance"] == "Non-Compliant").sum())

    def pct(n):
        return f"{(n / total_cms * 100):.0f}%" if total_cms else "0%"

    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card total">
            <div class="kpi-label">CMs Tracked</div>
            <div class="kpi-value">{total_cms}</div>
            <div class="kpi-sub">across all territories</div>
        </div>
        <div class="kpi-card full">
            <div class="kpi-label">Fully Compliant</div>
            <div class="kpi-value">{full_count}</div>
            <div class="kpi-sub">{pct(full_count)} of CMs</div>
        </div>
        <div class="kpi-card partial">
            <div class="kpi-label">Partially Compliant</div>
            <div class="kpi-value">{partial_count}</div>
            <div class="kpi-sub">{pct(partial_count)} of CMs</div>
        </div>
        <div class="kpi-card non">
            <div class="kpi-label">Non-Compliant</div>
            <div class="kpi-value">{non_count}</div>
            <div class="kpi-sub">{pct(non_count)} of CMs</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="checkpoint-divider"></div>', unsafe_allow_html=True)

    # ============================================================
    # Table preview
    # ============================================================

    st.markdown('<div class="section-label">Summary — one row per CM, with DM-wise subtotals</div>' if has_dm
                else '<div class="section-label">Summary — one row per CM</div>', unsafe_allow_html=True)

    def style_status(val):
        if val in ("Met", "Fully Compliant"):
            return "background-color:#E1F5F1; color:#127365; font-weight:600;"
        if val == "Partially Compliant":
            return "background-color:#FCF1DF; color:#9C6A17; font-weight:600;"
        if val in ("Not Met", "Non-Compliant"):
            return "background-color:#FBE7E4; color:#B03A2E; font-weight:600;"
        if val == DISTANCE_FLAG_LABEL:
            return "background-color:#FBE7E4; color:#B03A2E; font-weight:600;"
        return ""

    def style_total_rows(row):
        is_total = isinstance(row["S.No"], str) and row["S.No"].endswith("Total")
        if is_total:
            fill = "background-color:#FFF2CC; font-weight:700;" if row["S.No"] == "Grand Total" \
                else "background-color:#FDEBD0; font-weight:700;"
            return [fill] * len(row)
        return [""] * len(row)

    try:
        styled = summary_df.style.applymap(
            style_status,
            subset=[
                "FH Status",
                "SH Status",
                "Overall Compliance",
                "FH Distance (km)",
                "SH Distance (km)",
            ]
        )
        if has_dm:
            styled = styled.apply(style_total_rows, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True)
    except Exception:
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # ============================================================
    # Location-wise CRM report (Custom range only)
    # ============================================================

    if location_df is not None:
        st.markdown('<div class="checkpoint-divider"></div>', unsafe_allow_html=True)
        st.markdown('<div class="section-label">Location Summary — one row per location visited, with DM-wise subtotals</div>' if has_dm
                    else '<div class="section-label">Location Summary — one row per location visited</div>', unsafe_allow_html=True)
        st.caption(
            "One row per location visited (FH and SH visits combined). "
            "Total Leads are counted where Loan Requirement - Yes/No = \"Yes\". "
            "Total Logins and Total Non-Compliances are per-CM figures "
            "that repeat across the CM's location rows."
        )

        def style_non_compliance(val):
            if isinstance(val, (int, float)) and val > 0:
                return "background-color:#FBE7E4; color:#B03A2E; font-weight:600;"
            return ""

        def style_location_total_rows(row):
            is_total = isinstance(row["S.No"], str) and row["S.No"].endswith("Total")
            if is_total:
                fill = "background-color:#FFF2CC; font-weight:700;" if row["S.No"] == "Grand Total" \
                    else "background-color:#FDEBD0; font-weight:700;"
                return [fill] * len(row)
            return [""] * len(row)

        try:
            styled_location = location_df.style.applymap(
                style_non_compliance,
                subset=["Total Non-Compliances"]
            )
            if has_dm:
                styled_location = styled_location.apply(style_location_total_rows, axis=1)
            st.dataframe(styled_location, use_container_width=True, hide_index=True)
        except Exception:
            st.dataframe(location_df, use_container_width=True, hide_index=True)

    # ============================================================
    # Write formatted Excel to memory (single source of truth for download)
    # ============================================================

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        if location_df is not None:
            location_df.to_excel(writer, index=False, sheet_name="Location Summary")

    output.seek(0)

    # Reload the in-memory workbook to apply formatting
    wb = load_workbook(output)
    ws = wb["Summary"]

    # ==========================
    # Colors
    # ==========================

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")   # Dark Blue
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    dm_total_fill = PatternFill(fill_type="solid", fgColor="FFC000")   # Orange — DM subtotal rows
    grand_total_fill = PatternFill(fill_type="solid", fgColor="FAC090")  # Peach/gold — Grand Total row

    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    thin = Side(border_style="thin", color="000000")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    def format_sheet(ws):
        """Applies the shared header/border/autofit formatting to any sheet."""
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = border

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.border = border
                c.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 3, 40)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 35

    format_sheet(ws)

    # ==========================
    # Column Numbers
    # ==========================

    headers = {}
    for cell in ws[1]:
        headers[cell.value] = cell.column

    fh_status_col = headers["FH Status"]
    sh_status_col = headers["SH Status"]
    overall_col = headers["Overall Compliance"]
    fh_distance_col = headers["FH Distance (km)"]
    sh_distance_col = headers["SH Distance (km)"]
    sno_col = headers["S.No"]

    # ==========================
    # Row Formatting (Summary sheet conditional colors)
    # ==========================

    for row in range(2, ws.max_row + 1):

        sno_val = ws.cell(row=row, column=sno_col).value
        is_dm_total = isinstance(sno_val, str) and sno_val.startswith("DM-") and sno_val.endswith("Total")
        is_grand_total = sno_val == "Grand Total"

        if is_dm_total or is_grand_total:
            fill = grand_total_fill if is_grand_total else dm_total_fill
            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = bold_font
                c.border = border
            # Skip the per-status conditional coloring below for total rows
            continue

        c = ws.cell(row=row, column=fh_status_col)
        c.fill = green_fill if c.value == "Met" else red_fill

        c = ws.cell(row=row, column=sh_status_col)
        c.fill = green_fill if c.value == "Met" else red_fill

        c = ws.cell(row=row, column=overall_col)
        if c.value == "Fully Compliant":
            c.fill = green_fill
        elif c.value == "Partially Compliant":
            c.fill = yellow_fill
        elif c.value == "Non-Compliant":
            c.fill = red_fill

        if "Total Non-Compliances" in headers:
            c = ws.cell(row=row, column=headers["Total Non-Compliances"])
            if isinstance(c.value, (int, float)) and c.value > 0:
                c.fill = red_fill

        c = ws.cell(row=row, column=fh_distance_col)
        if c.value == DISTANCE_FLAG_LABEL:
            c.fill = red_fill

        c = ws.cell(row=row, column=sh_distance_col)
        if c.value == DISTANCE_FLAG_LABEL:
            c.fill = red_fill

    # ==========================
    # Location Summary sheet formatting (Custom range only)
    # ==========================

    if location_df is not None and "Location Summary" in wb.sheetnames:
        ws_loc = wb["Location Summary"]
        format_sheet(ws_loc)

        loc_headers = {}
        for cell in ws_loc[1]:
            loc_headers[cell.value] = cell.column

        non_compliance_col = loc_headers.get("Total Non-Compliances")
        loc_sno_col = loc_headers.get("S.No")

        # Merge CM-level (and DM-level) columns down each CM's location block.
        merge_cols = [
            loc_headers.get(name)
            for name in (
                "S.No", "DM Name", "Territory", "Cluster", "CM Name",
                "Total Logins", "Total Non-Compliances",
            )
            if loc_headers.get(name)
        ]

        current_row = 2
        for group_size in location_merge_sizes:
            end_row = current_row + group_size - 1

            sno_val = ws_loc.cell(row=current_row, column=loc_sno_col).value if loc_sno_col else None
            is_dm_total = isinstance(sno_val, str) and sno_val.startswith("DM-") and sno_val.endswith("Total")
            is_grand_total = sno_val == "Grand Total"

            if is_dm_total or is_grand_total:
                fill = grand_total_fill if is_grand_total else dm_total_fill
                for col in range(1, ws_loc.max_column + 1):
                    c = ws_loc.cell(row=current_row, column=col)
                    c.fill = fill
                    c.font = bold_font
                    c.border = border
                current_row = end_row + 1
                continue

            if non_compliance_col:
                top_cell = ws_loc.cell(row=current_row, column=non_compliance_col)
                if isinstance(top_cell.value, (int, float)) and top_cell.value > 0:
                    for r in range(current_row, end_row + 1):
                        ws_loc.cell(row=r, column=non_compliance_col).fill = red_fill

            if group_size > 1:
                for col in merge_cols:
                    ws_loc.merge_cells(
                        start_row=current_row, start_column=col,
                        end_row=end_row, end_column=col
                    )
                    top_cell = ws_loc.cell(row=current_row, column=col)
                    top_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

            current_row = end_row + 1

    # ==========================
    # Save formatted workbook back to memory for download
    # ==========================

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.markdown('<div class="checkpoint-divider"></div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="card">
        <div class="section-label">Export</div>
        <p style="color:var(--muted); margin:0 0 1rem 0;">
        Download the same summary as a formatted, ready-to-share Excel report.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        "Download summary report (.xlsx)",
        data=final_output,
        file_name="Summary_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )